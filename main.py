from jinja2 import Environment, FileSystemLoader
from ipaddress import IPv4Network
import ipaddress
import openpyxl

DEBUG = True
wb_obj = ""


def main():

    input_filename = "B2B VPN Jinja Templates.xlsx"
    var_list = ""
    xls_dict = ""
    xls_template_list = ""

    # If input file ends in 'xlsx', execute subroutine for xls parser to get variables and any templates
    if right(input_filename, 4) == "xlsx" or right(input_filename, 3) == "xls":
        xls_data_list, xls_template_list, xls_device_data = parse_xls_data(input_filename)
        #var_list = xls_data_list
        xls_dict = xls_device_data

    # If var_list is populated, execute remainder of script
    if var_list:
    #     # Cycle through each entry in XLS list
    #     for x in range(0, len(var_list)):
    #         # Add current entry to dictionary
    #         variable_dict = var_list[x]
    #         # run JINJA template against dictionary
    #         b2b_vpn_output = run_jinja_template("b2b_vpn_asa_config_template.yml", variable_dict)
    #         print(b2b_vpn_output)
    #         write_file(b2b_vpn_output, "B2B VPN - ASA Config_" + variable_dict["CUSTOMER_NAME"] + ".txt")
    #         b2b_peer_review_output = run_jinja_template("b2b_vpn_peer_review_template.yml", variable_dict)
    #         print(b2b_peer_review_output)
    #         write_file(b2b_peer_review_output, "B2B VPN - ASA Config_" + variable_dict["CUSTOMER_NAME"] + "_PEER_REVIEW.txt")
        pass
    elif xls_dict:
        #
        #   Cycle through each instance in the variable list
        #
        for id in xls_dict:
            #
            # Cycle through each device in the list
            #
            for device in xls_dict[id]['devices']:
                # If device name is found, cycle through templates for the device.
                template_output = ""
                name_value = xls_dict[id]['devices'][device]['name']
                if name_value != "":
                    for template in xls_dict[id]['devices'][device]['templates']:

                        if template in xls_template_list:
                            template_output = template_output + run_jinja_template("temp_template_" + template + ".txt", xls_dict[id]['variables'])

                        else:
                            print("Template " + template + " was not found.  Ignoring this template for the output.")
                    print(template_output)
                    output_name = "ID-" + str(id) + "-" + xls_dict[id]['devices'][device]['filename']
                    write_file(template_output, output_name)


    else:
        print("\n \nScript aborted.  Please correct the issue and re-run script.")


def parse_xls_data(input_filename):
    global wb_obj
    # Open XLS file
    wb_obj = open_xls(input_filename)
    # Find tab names in XLS file
    data_tabs = find_tabs_in_xls("Data")
    template_tabs = find_tabs_in_xls("Template")
    device_tabs = find_tabs_in_xls("Devices")
    # Find ID Column
    id_column_dict = find_column(data_tabs, "ID")
    # Find first data column
    first_data_column_dict = find_column(data_tabs, "ID:5")
    # Get all variables and values from each of the Data tabs and compile into a list entry for each column of data
    variable_list, variable_dict = get_xls_variables(data_tabs, id_column_dict, first_data_column_dict)
    # Enhance CIDR format addresses
    variable_list = enhance_cidr(variable_list)
    variable_dict = enhance_cidr(variable_dict)
    # Get all template data from each of the Template tabs and compile into list of dictionaries for each template found
    template_data = get_xls_templates(template_tabs)
    device_data = get_xls_devices(device_tabs, variable_dict)
    return variable_list, template_data, device_data


def get_xls_devices(xls_tabs, input_dict):
    # This function will cycle through all 'Devices' tabs in the spreadsheet and merge any relevant data into variable
    # dictionary
    output_dict = input_dict

    id_columns = find_column(xls_tabs, "ID")
    # Find first data column
    first_data_column_dict = find_column(xls_tabs, "ID:5")

    for tab in xls_tabs:
        sheet_obj = wb_obj[tab]
        variable_column = id_columns[tab]
        first_data_column = first_data_column_dict[tab]
        variable_list = []

        # Get all variable names in a list to parse
        for row in range(2, sheet_obj.max_row):
            current_value = rw_cell(row, variable_column, tab)
            variable_list.append(current_value)

        for column in range(first_data_column, sheet_obj.max_column):
            current_cell = rw_cell(1, column, tab)
            if current_cell == "" or current_cell is None:
                pass
            else:
                current_id = int(rw_cell(1, column, tab).split(":")[1])
                output_dict[current_id]['devices'] = {}
                #
                # Parse variables to look for devices that need to be created in our output dictionary for this column
                #
                for var in variable_list:
                    if left(var, 3) == "DN:":
                        var = right(var, len(var) - 3)
                        device_id = var.split(":")[0]
                        if device_id in output_dict[current_id]['devices']:
                            continue
                        else:
                            output_dict[current_id]['devices'][device_id] = {}
                            output_dict[current_id]['devices'][device_id]['name'] = ""
                            output_dict[current_id]['devices'][device_id]['filename'] = ""
                            output_dict[current_id]['devices'][device_id]['templates'] = []
                #
                # Parse variable column to determine where to place data in dictionary if data is found in current cell
                #
                for current_row in range(2, sheet_obj.max_column):
                    cell_data = rw_cell(current_row, column, tab)
                    variable_name = rw_cell(current_row, variable_column, tab)

                    if cell_data == "" or cell_data is None:
                        pass
                    else:
                        if left(variable_name, 2) == "DN":
                            device_id, device_name = split_variable(variable_name)
                            output_dict[current_id]['devices'][device_id]['name'] = cell_data
                        elif left(variable_name, 2) == "DT":
                            device_id, device_name = split_variable(variable_name)
                            output_dict[current_id]['devices'][device_id]['templates'].append(cell_data)
                        elif left(variable_name, 2) == "DF":
                            device_id, device_name = split_variable(variable_name)
                            output_dict[current_id]['devices'][device_id]['filename'] = cell_data
    return output_dict


def split_variable(input_name):
    short_name = right(input_name, len(input_name) - 3)
    first_part = short_name.split(":")[0]
    second_part = short_name.split(":")[1]
    return first_part, second_part


def get_xls_templates(template_tabs):
    # This function is used to find all the columns that have template data in them, grab the template data, and
    # return as a list to the higher function.
    template_list = {}
    for tab in template_tabs:
        sheet_obj = wb_obj[tab]
        temp_template_dict = {}

        for current_column in range(1, sheet_obj.max_column + 1):
            header_value = rw_cell(1, current_column, tab)
            if header_value is None:
                header_value = ""
            if left(header_value, 2) == "T:":
                template_name = header_value.split(":")[1]
                temp_template_dict[template_name] = {}

                row_data = get_xls_template_row(tab, current_column, 3, template_name)
                temp_template_dict[template_name]['template'] = ""
                temp_template_dict[template_name]['template'] = row_data
                template_list.update(temp_template_dict)
    return template_list


def get_xls_template_row(template_tab, column, max_blanks, template_name):
    # This function is used to find all the data in a row up to the max number of blank lines
    sheet_obj = wb_obj[template_tab]
    row_data = ""
    # Track consecutive blank rows
    consecutive_blanks = 0

    for row in range(2, sheet_obj.max_row + 1):
        cell_value = rw_cell(row, column, template_tab)
        if cell_value is None:
            consecutive_blanks += 1
            cell_value = ""
        else:
            consecutive_blanks = 0
        if consecutive_blanks == max_blanks:
            break
        else:
            if row_data == "":
                row_data += cell_value
            else:
                row_data += "\n" + cell_value

    file = open("temp_template_" + template_name + ".txt", "w")
    file.write(row_data)
    file.close()

    return row_data


def find_tabs_in_xls(find_name):
    tab_names = wb_obj.sheetnames
    matched_tabs = []

    for tab in tab_names:
        find_name_length = len(find_name)
        tab_name_split = left(tab, find_name_length)
        if find_name == tab_name_split:
            if DEBUG:
                print("Found tab named " + tab)
            matched_tabs.append(tab)
    return matched_tabs


def write_file(input_string, filename):
    file_object = open("output/" + filename, "w+")
    file_object.write(input_string)
    file_object.close()


def enhance_cidr(data_input):
    new_data_list = []
    my_dict = {}

    if isinstance(data_input, list):
        for x in data_input:
            new_dict = {}
            for key, value in x.items():
                try:
                    if is_cidr_format(value):
                        cidr_dict = convert_cidr(value, key)
                        new_dict[key] = value
                        new_dict.update(cidr_dict)

                    else:
                        new_dict[key] = value
                except:
                    new_dict[key] = value

            new_data_list.append(new_dict)
        return new_data_list
    elif isinstance(data_input, dict):
        for x in list(data_input):
            my_dict[x] = {}
            my_dict[x]['variables'] = data_input[x]['variables']
            cidr_dict = {}
            for key in data_input[x]['variables']:
                current_value = data_input[x]['variables'][key]
            #for key, value in data_input[x]['variables'].items():
                try:
                    if is_cidr_format(current_value):
                        cidr_dict.update(convert_cidr(current_value, key))
                        #cidr_dict = convert_cidr(current_value, key)
                        #my_dict[x]['variables'].update(cidr_dict)
                except:
                    pass
            my_dict[x]['variables'].update(cidr_dict)
        return my_dict


def get_xls_variables(tab_list, id_column_dict, first_data_column_dict):
    # Create local list to pass back from function
    my_list = []
    my_dict = {}
    # Cycle through each tab in the tab_list
    for tab in tab_list:
        sheet_obj = wb_obj[tab]
        for c in range(first_data_column_dict[tab], sheet_obj.max_column + 1):
            if rw_cell(1, c, tab):
                if c not in my_dict:
                    my_dict[c] = {}
                    my_dict[c]['variables'] = {}
                for i in range(1,sheet_obj.max_row + 1):
                    variable_name = rw_cell(i, id_column_dict[tab], tab)
                    if variable_name:
                        if left(variable_name, 3) == "DD:":
                            cell_value = rw_cell(i, c, tab)
                            if cell_value:
                                my_dict[c]['variables'][variable_name.split(":")[1]] = cell_value

    # Convert nested dictionary to flat list
    my_list = flatten_dict(my_dict)
    return my_list, my_dict


def run_jinja_template(jinja_file, variable_dict):
    # This line uses the current directory
    file_loader = FileSystemLoader('.')
    # Load the enviroment
    env = Environment(loader=file_loader)
    # Open single Jinja2 file and load as template
    template = env.get_template(jinja_file)

    # Parse template with dictionary created
    output = template.render(variable_dict)
    return output


def flatten_dict(dict):
    # This function flattens a nested dictionary into a single list
    result = []
    for key in dict:
        new_dict = {}
        new_dict['ID'] = key
        for subkey in dict[key]:
            new_dict[subkey] = dict[key][subkey]
        result.append(new_dict)
    return result


def open_xls(xls_input_file):
    workbook = openpyxl.load_workbook(xls_input_file, data_only=True)
    return workbook


def find_column(tab_list, value_to_look_for):
    column_dict = {}
    for tab in tab_list:
        sheet_obj = wb_obj[tab]
        for i in range(1, sheet_obj.max_column + 1):
            if rw_cell(1, i, tab) == value_to_look_for:
                column_dict[tab] = i

    return column_dict


def is_cidr_format(cidr_address):
    if cidr_address.find("/") > 0:
        split = cidr_address.split("/")
        try:
            result = ipaddress.ip_address(split[0])
            if 0 < int(split[1]) <= 32:
                #print("The original string was " + cidr_address + ".  IP address is: " + str(
                #    result) + " and subnet length is " + split[1])
                return True
            else:
                pass
                #print("The string " + cidr_address + " contains an invalid subnet length.  This will result in an incomplete configuration and cause the deployment to fail.  Please correct this error and resubmit.")
        except:
            # print ("The string " + cidr_address + " contains an invalid IP Address.  This will result in an incomplete configuration and cause the deployment to fail.  Please correct this error and resubmit.")
            return False
    else:
        return False


def convert_cidr(input_cidr_dictionary, output_name):
    # This module will take an input and parse it using the ipaddress module
    net = IPv4Network(input_cidr_dictionary)
    cidr_dict = {output_name + "_IP": str(net.network_address), output_name + "_NETMASK": str(net.netmask),
                 output_name + "_WILDCARD": str(net.hostmask), output_name + "_LENGTH": int(net.prefixlen)}
    return cidr_dict


def rw_cell(row, column, sheet, write=False, value=""):
    global wb_obj
    sheet = wb_obj[sheet]
    if sheet == "":
        return
    if write is False:
        value = sheet.cell(row=row, column=column).value
        return value
    elif write is True:
        sheet.cell(row=row, column=column).value = value


def left(input_string, amount):
    return input_string[:amount]


def right(input_string, amount):
    return input_string[-amount:]


def mid(input_string, offset, amount):
    return input_string[offset:offset+amount]


main()
